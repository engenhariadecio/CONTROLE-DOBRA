"""
Controle de Produtividade — DOBRA (e demais setores da conformação)
====================================================================
Fluxo do operador:
    Escolhe a máquina  ->  Inicia (informa OP/operador)  ->  Produz
    ->  Pausa/Retoma nos intervalos  ->  Finaliza (qtd + refugo)

Recursos:
- Grade de máquinas por setor (Dobra em foco; serve p/ Corte, Estamparia,
  Solda, Acabamento — basta cadastrar as máquinas do setor).
- Cronômetro por máquina com PAUSA que desconta do tempo produtivo
  (café, almoço, ginástica, banheiro, manutenção, falta de material...).
- Painel Admin: cadastro de máquinas, usuários, setores e motivos de pausa.
- Painel Gerencial (dashboard) com indicadores e gráficos.
- Exportação Excel dos apontamentos.

Banco: PostgreSQL (produção, ex.: Railway via DATABASE_URL).
       Sem DATABASE_URL -> SQLite local (dobra.db) para rodar na sua máquina.
"""
import os
import io
import csv
import json
import threading
from datetime import datetime, timedelta, date as ddate
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, send_file
)
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import (
    create_engine, Column, Integer, String, DateTime, Float, Text,
    Boolean, inspect, text
)
from sqlalchemy.orm import declarative_base, sessionmaker, scoped_session
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'troque-esta-chave-em-producao')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_REFRESH_EACH_REQUEST'] = True

# Fuso local (Brasil UTC-3). Datas são gravadas em UTC e exibidas no local.
FUSO_LOCAL_HORAS = 3

# ─────────────────────────────────────────────────────────────────────────────
# Banco de dados
# ─────────────────────────────────────────────────────────────────────────────
def _database_url():
    url = os.environ.get('DATABASE_URL', '').strip()
    if url.startswith('postgres://'):
        # SQLAlchemy exige o prefixo postgresql://
        url = url.replace('postgres://', 'postgresql://', 1)
    if not url:
        base = os.path.dirname(os.path.abspath(__file__))
        url = 'sqlite:///' + os.path.join(base, 'dobra.db')
    return url


DB_URL = _database_url()
_engine_kwargs = {'pool_pre_ping': True, 'future': True}
if DB_URL.startswith('sqlite'):
    _engine_kwargs['connect_args'] = {'check_same_thread': False}

engine = create_engine(DB_URL, **_engine_kwargs)
Session = scoped_session(sessionmaker(bind=engine, autoflush=False, future=True))
Base = declarative_base()


# ─────────────────────────────────────────────────────────────────────────────
# Lista mestra do SAP (mesma abordagem do sistema Banho): arquivo carregado em
# memória. Usada para "bipar" a OP e puxar material, texto breve e quantidade.
#   Arquivos aceitos na raiz: lista_mestra.xlsx / .csv / .txt (ou o exemplo).
# ─────────────────────────────────────────────────────────────────────────────
_lista_lock = threading.Lock()
_lista_por_ordem = {}
_lista_status = {'carregada': False, 'total': 0, 'erro': None, 'arquivo': ''}

LISTA_MESTRA_ARQUIVOS = [
    'lista_mestra.xlsx', 'lista_mestra.csv', 'lista_mestra.txt',
    'exemplo_lista_mestra_sap.txt',
]


def _norm_ordem(v):
    """Normaliza a OP bipada: remove '.0', e se o código de barras vier com mais
    de 8 dígitos, descarta 4 prefixos e 4 sufixos (padrão do leitor do Banho)."""
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    digitos = ''.join(c for c in s if c.isdigit())
    if len(digitos) > 8:
        digitos = digitos[4:-4]
    return digitos if digitos else s


def _achar_arquivo_mestre():
    base = os.path.dirname(os.path.abspath(__file__))
    for nome in LISTA_MESTRA_ARQUIVOS:
        caminho = os.path.join(base, nome)
        if os.path.isfile(caminho):
            return caminho
    return None


def _achar_colunas(linhas):
    def norm(s):
        return str(s).strip().lower() if s is not None else ''
    for i, row in enumerate(linhas[:10]):
        if not row:
            continue
        idx = {}
        for j, nome in enumerate(norm(c) for c in row):
            if nome == 'ordem' and 'ordem' not in idx:
                idx['ordem'] = j
            elif nome == 'material' and 'material' not in idx:
                idx['material'] = j
            elif 'texto breve' in nome and 'texto' not in idx:
                idx['texto'] = j
            elif ('quantidade da ordem' in nome or nome == 'quantidade total'
                  or nome == 'quantidade') and 'qtd' not in idx:
                idx['qtd'] = j
        if 'ordem' in idx and 'material' in idx:
            return i, idx
    return None


def _parsear_linhas_mestre(linhas):
    achado = _achar_colunas(linhas)
    if achado:
        cab, col = achado
        i_ordem, i_mat = col.get('ordem', 0), col.get('material', 1)
        i_texto, i_qtd = col.get('texto'), col.get('qtd')
        inicio = cab + 1
    else:
        i_ordem, i_mat, i_texto, i_qtd, inicio = 0, 2, 3, 4, 0

    def val(row, idx):
        if idx is None or idx >= len(row) or row[idx] is None:
            return ''
        return str(row[idx]).strip()

    por_ordem = {}
    for row in linhas[inicio:]:
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        ordem = _norm_ordem(row[i_ordem]) if i_ordem < len(row) and row[i_ordem] is not None else ''
        if not ordem or not ordem.replace('.', '').isdigit():
            continue
        q = val(row, i_qtd)
        try:
            qtd = int(float(q)) if q else 0
        except (ValueError, TypeError):
            qtd = 0
        por_ordem[ordem] = {'ordem': ordem, 'material': val(row, i_mat),
                            'texto_breve': val(row, i_texto), 'quantidade': qtd}
    return por_ordem


def carregar_lista_mestre():
    global _lista_por_ordem, _lista_status
    caminho = _achar_arquivo_mestre()
    if not caminho:
        with _lista_lock:
            _lista_status = {'carregada': False, 'total': 0, 'arquivo': '',
                             'erro': 'Arquivo lista_mestra.xlsx/.csv/.txt não encontrado.'}
        return
    try:
        nome = caminho.lower()
        linhas = []
        if nome.endswith('.csv') or nome.endswith('.txt'):
            with open(caminho, encoding='utf-8-sig', errors='replace') as f:
                raw = f.read()
            sep = '\t' if raw.count('\t') > raw.count(';') and raw.count('\t') > raw.count(',') \
                else (';' if raw.count(';') > raw.count(',') else ',')
            linhas = list(csv.reader(io.StringIO(raw), delimiter=sep))
        else:
            from openpyxl import load_workbook as _lw
            wb = _lw(caminho, read_only=True, data_only=True)
            for row in wb.active.iter_rows(values_only=True):
                linhas.append(list(row))
        por_ordem = _parsear_linhas_mestre(linhas)
        with _lista_lock:
            _lista_por_ordem = por_ordem
            _lista_status = {'carregada': True, 'total': len(por_ordem),
                             'arquivo': os.path.basename(caminho), 'erro': None}
        print(f'[lista_mestra] Carregada: {len(por_ordem)} ordens de "{os.path.basename(caminho)}".')
    except Exception as e:  # noqa: BLE001
        with _lista_lock:
            _lista_status = {'carregada': False, 'total': 0, 'arquivo': '', 'erro': str(e)}
        print(f'[lista_mestra] ERRO ao carregar: {e}')


# ─────────────────────────────────────────────────────────────────────────────
# Modelos
# ─────────────────────────────────────────────────────────────────────────────
class Usuario(Base):
    __tablename__ = 'usuarios'
    id = Column(Integer, primary_key=True)
    login = Column(String(50), unique=True, nullable=False)
    nome = Column(String(120), nullable=False)
    matricula = Column(String(50), default='')
    senha_hash = Column(String(255), nullable=False)
    perfil = Column(String(20), nullable=False)        # perfil principal
    perfis = Column(Text, default='')                  # "operador,gerencia,admin"
    setor = Column(String(40), default='')             # área do usuário ('' = todos os setores)
    ativo = Column(Boolean, default=True)

    def acessos(self):
        lst = [p.strip() for p in (self.perfis or '').split(',') if p.strip()]
        if not lst and self.perfil:
            lst = [self.perfil]
        vistos = []
        for p in lst:
            if p not in vistos:
                vistos.append(p)
        return vistos

    def to_dict(self):
        return {'id': self.id, 'login': self.login, 'nome': self.nome,
                'matricula': self.matricula or '', 'perfil': self.perfil,
                'perfis': self.acessos(), 'setor': self.setor or '',
                'ativo': bool(self.ativo)}


class Maquina(Base):
    __tablename__ = 'maquinas'
    id = Column(Integer, primary_key=True)
    nome = Column(String(120), nullable=False)
    codigo = Column(String(40), default='')            # etiqueta curta (ex.: DOB-01)
    setor = Column(String(40), default='Dobra')
    ativa = Column(Boolean, default=True)
    ordem = Column(Integer, default=0)                 # ordenação na grade
    meta_pph = Column(Float, default=0)                # meta de peças/hora (capacidade nominal p/ OEE)
    criado_em = Column(DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {'id': self.id, 'nome': self.nome, 'codigo': self.codigo or '',
                'setor': self.setor or '', 'ativa': bool(self.ativa),
                'ordem': self.ordem or 0,
                'meta_pph': round(self.meta_pph or 0, 2)}


def calc_oee(prod_seg, pausa_seg, qtd, refugo, meta_pph):
    """Calcula os pilares do OEE a partir de somas brutas.

    Disponibilidade = tempo produtivo / (produtivo + pausa)
    Desempenho      = peças produzidas / (meta_pph * horas produtivas)   [cap 100%]
    Qualidade       = (produzidas - refugo) / produzidas
    OEE             = Disponibilidade x Desempenho x Qualidade

    Cada pilar é retornado como fração 0..1, ou None quando não há base
    (ex.: sem meta de peças/hora cadastrada -> desempenho e OEE = None).
    """
    prod_seg = max(0.0, float(prod_seg or 0.0))
    pausa_seg = max(0.0, float(pausa_seg or 0.0))
    qtd = max(0, int(qtd or 0))
    refugo = max(0, int(refugo or 0))
    meta = max(0.0, float(meta_pph or 0.0))

    total = prod_seg + pausa_seg
    disp = (prod_seg / total) if total > 0 else None

    teorico = meta * (prod_seg / 3600.0)
    if teorico > 0:
        desemp = min(qtd / teorico, 1.0)
    else:
        desemp = None

    qual = ((qtd - refugo) / qtd) if qtd > 0 else None

    if None in (disp, desemp, qual):
        oee = None
    else:
        oee = disp * desemp * qual

    return {'disponibilidade': disp, 'desempenho': desemp,
            'qualidade': qual, 'oee': oee}


def _pct(x):
    """Fração 0..1 -> percentual (1 casa), ou None."""
    return round(x * 100, 1) if x is not None else None


class Apontamento(Base):
    __tablename__ = 'apontamentos'
    id = Column(Integer, primary_key=True)
    estado = Column(String(20), nullable=False, index=True)  # produzindo|pausado|finalizado

    maquina_id = Column(Integer, index=True)
    maquina_nome = Column(String(120), default='')
    setor = Column(String(40), default='Dobra')

    operador_nome = Column(String(120), default='')
    operador_matricula = Column(String(50), default='')

    op = Column(String(60), default='')                # ordem de produção
    codigo = Column(String(60), default='')            # material/código
    descricao = Column(String(255), default='')
    quantidade_prevista = Column(Integer, default=0)
    quantidade_produzida = Column(Integer, default=0)
    refugo = Column(Integer, default=0)
    meta_pph = Column(Float, default=0)                # meta peças/hora vigente no apontamento (p/ OEE)

    inicio = Column(DateTime)
    fim = Column(DateTime)
    producao_seg = Column(Float, default=0)            # tempo produtivo final (s)

    pausado = Column(Integer, default=0)
    pausa_inicio = Column(DateTime)
    pausa_motivo = Column(String(80), default='')
    pausa_acumulada_seg = Column(Integer, default=0)
    pausas_json = Column(Text, default='')             # [{ini,fim,motivo,seg}]

    observacao = Column(Text, default='')
    criado_em = Column(DateTime, default=datetime.utcnow)

    # ── tempo ──
    def _pausas(self):
        try:
            return json.loads(self.pausas_json) if self.pausas_json else []
        except (ValueError, TypeError):
            return []

    def pausa_total_seg(self, agora=None):
        base = self.pausa_acumulada_seg or 0
        if self.estado == 'pausado' and self.pausa_inicio and not self.fim:
            agora = agora or datetime.utcnow()
            base += (agora - self.pausa_inicio).total_seconds()
        return max(0, base)

    def produtivo_seg(self, agora=None):
        if not self.inicio:
            return 0.0
        agora = agora or datetime.utcnow()
        fim = self.fim or agora
        bruto = (fim - self.inicio).total_seconds()
        return max(0.0, bruto - self.pausa_total_seg(agora))

    def meta_efetiva(self, meta_padrao=0):
        """Meta de peças/hora vigente: a do apontamento, ou o padrão global."""
        return float(self.meta_pph or 0) or float(meta_padrao or 0)

    def oee(self, agora=None, meta_padrao=0):
        agora = agora or datetime.utcnow()
        prod = self.producao_seg if self.estado == 'finalizado' else self.produtivo_seg(agora)
        return calc_oee(prod, self.pausa_total_seg(agora),
                        self.quantidade_produzida or 0, self.refugo or 0,
                        self.meta_efetiva(meta_padrao))

    def to_dict(self, agora=None, meta_padrao=0):
        agora = agora or datetime.utcnow()

        def fmt(dt):
            return (dt - timedelta(hours=FUSO_LOCAL_HORAS)).strftime('%d/%m/%Y %H:%M:%S') if dt else ''

        def iso(dt):
            return dt.isoformat() + 'Z' if dt else ''

        prod = self.producao_seg if self.estado == 'finalizado' else self.produtivo_seg(agora)
        pausa = self.pausa_total_seg(agora)
        pausas = self._pausas()
        qprod = self.quantidade_produzida or 0
        pph = round(qprod / (prod / 3600), 1) if prod > 0 and qprod else 0.0
        meta = self.meta_efetiva(meta_padrao)
        o = calc_oee(prod, pausa, qprod, self.refugo or 0, meta)
        inicio_local = (self.inicio - timedelta(hours=FUSO_LOCAL_HORAS)) if self.inicio else None
        return {
            'id': self.id, 'estado': self.estado,
            'maquina_id': self.maquina_id, 'maquina_nome': self.maquina_nome,
            'setor': self.setor,
            'operador_nome': self.operador_nome, 'operador_matricula': self.operador_matricula or '',
            'op': self.op or '', 'codigo': self.codigo or '', 'descricao': self.descricao or '',
            'quantidade_prevista': self.quantidade_prevista or 0,
            'quantidade_produzida': qprod, 'refugo': self.refugo or 0,
            'inicio': fmt(self.inicio), 'inicio_iso': iso(self.inicio),
            'turno': turno_de(inicio_local),
            'fim': fmt(self.fim), 'fim_iso': iso(self.fim),
            'producao_seg': round(prod, 1), 'producao_min': round(prod / 60, 1),
            'pausa_seg': round(pausa, 1), 'pausa_min': round(pausa / 60, 1),
            'pausado': 1 if self.estado == 'pausado' else 0,
            'pausa_motivo': self.pausa_motivo or '',
            'pausa_inicio_iso': iso(self.pausa_inicio),
            'n_pausas': len(pausas), 'pausas': pausas,
            'pph': pph,
            'meta_pph': round(meta, 2),
            'disponibilidade': _pct(o['disponibilidade']),
            'desempenho': _pct(o['desempenho']),
            'qualidade': _pct(o['qualidade']),
            'oee': _pct(o['oee']),
            'observacao': self.observacao or '',
            'criado_em': fmt(self.criado_em),
        }


class Config(Base):
    __tablename__ = 'config'
    id = Column(Integer, primary_key=True)
    chave = Column(String(60), unique=True, nullable=False)
    valor = Column(Text, default='')


class ApontamentoLog(Base):
    """Auditoria de alterações/exclusões feitas pelo admin."""
    __tablename__ = 'apontamento_logs'
    id = Column(Integer, primary_key=True)
    apontamento_id = Column(Integer, index=True)
    quando = Column(DateTime, default=datetime.utcnow)
    usuario = Column(String(120), default='')
    acao = Column(String(30), default='editar')
    antes_json = Column(Text, default='')
    depois_json = Column(Text, default='')


# ─────────────────────────────────────────────────────────────────────────────
# Configurações do sistema (setores, motivos de pausa, setor ativo)
# ─────────────────────────────────────────────────────────────────────────────
SETORES_PADRAO = ['Dobra', 'Corte', 'Estamparia', 'Solda', 'Acabamento']
MOTIVOS_PAUSA_PADRAO = [
    'Almoço/Janta', 'Café', 'Laboral', 'Banheiro',
    'Manutenção', 'Falta de material', 'Setup / troca de ferramenta', 'Reunião',
]


def cfg_get(chave, padrao=None):
    db = Session()
    try:
        row = db.query(Config).filter_by(chave=chave).first()
        if row and row.valor:
            try:
                return json.loads(row.valor)
            except (ValueError, TypeError):
                return row.valor
        return padrao
    finally:
        db.close()


def cfg_set(chave, valor):
    db = Session()
    try:
        row = db.query(Config).filter_by(chave=chave).first()
        if not row:
            row = Config(chave=chave)
            db.add(row)
        row.valor = json.dumps(valor, ensure_ascii=False)
        db.commit()
    finally:
        db.close()


def get_setores():
    return cfg_get('setores', SETORES_PADRAO) or SETORES_PADRAO


def get_setor_ativo():
    return cfg_get('setor_ativo', 'Dobra') or 'Dobra'


def get_motivos_pausa():
    return cfg_get('motivos_pausa', MOTIVOS_PAUSA_PADRAO) or MOTIVOS_PAUSA_PADRAO


META_PPH_PADRAO = 0        # peças/hora usadas quando a máquina não tem meta própria (0 = sem meta)
META_OEE_PADRAO = 75       # meta de OEE (%) exibida como referência nos indicadores


def get_meta_pph_padrao():
    try:
        return float(cfg_get('meta_pph_padrao', META_PPH_PADRAO) or 0)
    except (ValueError, TypeError):
        return 0.0


def get_meta_oee():
    try:
        return float(cfg_get('meta_oee', META_OEE_PADRAO) or 0)
    except (ValueError, TypeError):
        return float(META_OEE_PADRAO)


# Turnos (3 turnos, iguais aos do sistema Banho). Horários locais HH:MM.
TURNOS_PADRAO = [
    {'nome': '1º turno', 'inicio': '06:01', 'fim': '15:30'},
    {'nome': '2º turno', 'inicio': '15:31', 'fim': '00:00'},
    {'nome': '3º turno', 'inicio': '00:01', 'fim': '06:00'},
]


def get_turnos():
    t = cfg_get('turnos', TURNOS_PADRAO) or TURNOS_PADRAO
    # saneia
    out = []
    for x in t:
        try:
            out.append({'nome': str(x.get('nome') or '').strip() or 'Turno',
                        'inicio': str(x.get('inicio') or '00:00'),
                        'fim': str(x.get('fim') or '00:00')})
        except AttributeError:
            continue
    return out or TURNOS_PADRAO


def _min_do_dia(hhmm):
    try:
        h, m = str(hhmm).split(':')
        return int(h) * 60 + int(m)
    except (ValueError, TypeError):
        return 0


def turno_de(dt_local):
    """Nome do turno de um datetime LOCAL, tratando turno que vira a meia-noite."""
    if not dt_local:
        return ''
    minutos = dt_local.hour * 60 + dt_local.minute
    # 00:00 exato conta como fim do dia (pertence ao turno que termina à meia-noite)
    mm = 1440 if minutos == 0 else minutos
    for t in get_turnos():
        ini = _min_do_dia(t['inicio'])
        fim = _min_do_dia(t['fim'])
        if fim == 0:
            fim = 24 * 60  # 00:00 = fim do dia
        if ini <= fim:
            if ini <= mm <= fim:
                return t['nome']
        else:  # vira a meia-noite (ex.: 22:00–06:00)
            if mm >= ini or mm <= fim:
                return t['nome']
    return ''


# ─────────────────────────────────────────────────────────────────────────────
# Bootstrap: cria tabelas e semeia dados iniciais
# ─────────────────────────────────────────────────────────────────────────────
def _garantir_colunas():
    """Migração leve: adiciona colunas novas se o banco for antigo (SQLite/PG)."""
    insp = inspect(engine)
    esperado = {
        'usuarios': {
            'matricula': "VARCHAR(50) DEFAULT ''",
            'perfis': "TEXT DEFAULT ''",
            'setor': "VARCHAR(40) DEFAULT ''",
            'ativo': 'BOOLEAN DEFAULT TRUE',
        },
        'maquinas': {
            'meta_pph': 'FLOAT DEFAULT 0',
        },
        'apontamentos': {
            'meta_pph': 'FLOAT DEFAULT 0',
        },
    }
    for tabela, cols in esperado.items():
        if tabela not in insp.get_table_names():
            continue
        existentes = {c['name'] for c in insp.get_columns(tabela)}
        for col, ddl in cols.items():
            if col not in existentes:
                try:
                    with engine.begin() as conn:
                        conn.execute(text(f'ALTER TABLE {tabela} ADD COLUMN {col} {ddl}'))
                except Exception as e:
                    print(f'[migracao] {tabela}.{col}: {e}')


def bootstrap():
    Base.metadata.create_all(engine)
    _garantir_colunas()
    db = Session()
    try:
        # Admin padrão
        if not db.query(Usuario).filter_by(login='admin').first():
            db.add(Usuario(
                login='admin', nome='Administrador',
                senha_hash=generate_password_hash('admin123'),
                perfil='admin', perfis='admin,gerencia,operador', ativo=True))
            print('[bootstrap] usuário admin criado (admin / admin123)')
        # Operador de exemplo
        if not db.query(Usuario).filter_by(login='operador').first():
            db.add(Usuario(
                login='operador', nome='Operador Dobra',
                senha_hash=generate_password_hash('123456'),
                perfil='operador', perfis='operador', setor='Dobra', ativo=True))
        # Máquinas de exemplo (Dobra)
        if db.query(Maquina).count() == 0:
            # Parque de máquinas de exemplo, cobrindo todos os setores da fábrica.
            # (código, nome, setor, meta peças/hora de referência)
            exemplos = [
                # Corte
                ('COR-PUN-01', 'Puncionadeira 01', 'Corte', 40),
                ('COR-PUN-02', 'Puncionadeira 02', 'Corte', 40),
                ('COR-LAS-01', 'Laser 01', 'Corte', 55),
                ('COR-LAS-02', 'Laser 02', 'Corte', 55),
                # Estamparia
                ('EST-FUR-01', 'Furadeira 01', 'Estamparia', 90),
                ('EST-PRE-01', 'Prensa 01', 'Estamparia', 120),
                ('EST-PRE-02', 'Prensa 02', 'Estamparia', 120),
                # Dobra
                ('DOB-01', 'Dobradeira 01', 'Dobra', 60),
                ('DOB-02', 'Dobradeira 02', 'Dobra', 60),
                ('DOB-03', 'Dobradeira 03', 'Dobra', 60),
                ('DOB-04', 'Dobradeira 04', 'Dobra', 60),
                # Solda
                ('SOL-PON-01', 'Solda Ponto 01', 'Solda', 75),
                ('SOL-PON-02', 'Solda Ponto 02', 'Solda', 75),
                ('SOL-01', 'Solda 01', 'Solda', 30),
                ('SOL-02', 'Solda 02', 'Solda', 30),
                # Acabamento
                ('ACA-POL-01', 'Polimento 01', 'Acabamento', 50),
                ('ACA-01', 'Acabamento 01', 'Acabamento', 50),
            ]
            for ordem, (cod, nome, setor, meta) in enumerate(exemplos, 1):
                db.add(Maquina(nome=nome, codigo=cod, setor=setor,
                               ativa=True, ordem=ordem, meta_pph=meta))
            print(f'[bootstrap] {len(exemplos)} máquinas de exemplo criadas em 5 setores')
        db.commit()
    finally:
        db.close()
    # Config padrão
    if cfg_get('setores') is None:
        cfg_set('setores', SETORES_PADRAO)
    if cfg_get('setor_ativo') is None:
        cfg_set('setor_ativo', 'Dobra')
    if cfg_get('motivos_pausa') is None:
        cfg_set('motivos_pausa', MOTIVOS_PAUSA_PADRAO)
    if cfg_get('meta_pph_padrao') is None:
        cfg_set('meta_pph_padrao', META_PPH_PADRAO)
    if cfg_get('meta_oee') is None:
        cfg_set('meta_oee', META_OEE_PADRAO)
    if cfg_get('turnos') is None:
        cfg_set('turnos', TURNOS_PADRAO)


with app.app_context():
    try:
        bootstrap()
    except Exception as e:
        print('[bootstrap] ERRO:', e)

try:
    carregar_lista_mestre()
except Exception as e:  # noqa: BLE001
    print('[lista_mestra] ERRO ao carregar no startup:', e)


@app.teardown_appcontext
def _remove_session(exc=None):
    Session.remove()


# ─────────────────────────────────────────────────────────────────────────────
# Autenticação
# ─────────────────────────────────────────────────────────────────────────────
def login_obrigatorio(fn):
    @wraps(fn)
    def wrap(*a, **k):
        if not session.get('usuario'):
            return redirect(url_for('login'))
        return fn(*a, **k)
    return wrap


def perfil_obrigatorio(*perfis):
    def deco(fn):
        @wraps(fn)
        def wrap(*a, **k):
            if not session.get('usuario'):
                return redirect(url_for('login'))
            acessos = session.get('perfis') or [session.get('perfil')]
            if 'admin' in acessos:
                return fn(*a, **k)
            if not any(p in acessos for p in perfis):
                return redirect(url_for('selecionar'))
            return fn(*a, **k)
        return wrap
    return deco


def api_login_obrigatorio(fn):
    @wraps(fn)
    def wrap(*a, **k):
        if not session.get('usuario'):
            return jsonify({'ok': False, 'erro': 'Sessão expirada. Entre novamente.'}), 401
        return fn(*a, **k)
    return wrap


@app.context_processor
def _inj():
    return {
        'nome': session.get('nome'),
        'setor_ativo': get_setor_ativo(),
        'usuario_setor': session.get('setor') or '',
    }


def setor_do_usuario():
    """Setor padrão do usuário logado: o setor cadastrado, senão o setor ativo global."""
    return (session.get('setor') or '').strip() or get_setor_ativo()


def _pode_ver_todos_setores():
    """Só admin ou usuário sem setor fixo pode ver 'Todos os setores'."""
    acessos = session.get('perfis') or [session.get('perfil')]
    return ('admin' in acessos) or not (session.get('setor') or '').strip()


def resolver_setor(arg):
    """Resolve o filtro de setor. Retorna (setor_ou_None, rotulo).
    setor None = todos os setores (só permitido a admin/geral)."""
    arg = (arg or '').strip()
    if arg == '__todos__':
        if _pode_ver_todos_setores():
            return None, 'Todos os setores'
        return setor_do_usuario(), setor_do_usuario()
    setor = arg or setor_do_usuario()
    return setor, setor


# ─────────────────────────────────────────────────────────────────────────────
# Rotas de sessão
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if not session.get('usuario'):
        return redirect(url_for('login'))
    return redirect(url_for('selecionar'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        login_ = (request.form.get('usuario') or '').strip()
        senha = request.form.get('senha') or ''
        db = Session()
        try:
            u = db.query(Usuario).filter_by(login=login_).first()
            if not u or not check_password_hash(u.senha_hash, senha):
                return render_template('login.html', erro='Usuário ou senha inválidos.')
            if not u.ativo:
                return render_template('login.html', erro='Usuário desativado. Fale com o admin.')
            session.permanent = True
            session['usuario'] = u.login
            session['nome'] = u.nome
            session['perfil'] = u.perfil
            session['perfis'] = u.acessos()
            session['matricula'] = u.matricula or ''
            session['setor'] = u.setor or ''
        finally:
            db.close()
        return redirect(url_for('selecionar'))
    if session.get('usuario'):
        return redirect(url_for('selecionar'))
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


PERFIL_DESTINO = {
    'operador': 'painel',
    'gerencia': 'dashboard',
    'admin': 'dashboard',
}


@app.route('/selecionar')
@login_obrigatorio
def selecionar():
    acessos = session.get('perfis') or [session.get('perfil')]
    # Se só tem um acesso, vai direto
    if len(acessos) == 1:
        destino = PERFIL_DESTINO.get(acessos[0], 'painel')
        return redirect(url_for(destino))
    return render_template('selecionar.html', acessos=acessos)


@app.route('/ir/<perfil>')
@login_obrigatorio
def ir(perfil):
    acessos = session.get('perfis') or [session.get('perfil')]
    if perfil not in acessos and 'admin' not in acessos:
        return redirect(url_for('selecionar'))
    destino = PERFIL_DESTINO.get(perfil, 'painel')
    return redirect(url_for(destino))


# ─────────────────────────────────────────────────────────────────────────────
# Páginas
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/painel')
@perfil_obrigatorio('operador')
def painel():
    acessos = session.get('perfis') or [session.get('perfil')]
    meu_setor = (session.get('setor') or '').strip()
    # Usuário sem setor fixo (ex.: admin) pode escolher qualquer setor.
    if meu_setor and 'admin' not in acessos:
        setores = [meu_setor]
    else:
        setores = get_setores()
    return render_template('painel.html', motivos=get_motivos_pausa(),
                           setores=setores, meu_setor=meu_setor or get_setor_ativo())


@app.route('/dashboard')
@app.route('/gerencia')
@perfil_obrigatorio('gerencia', 'admin')
def dashboard():
    acessos = session.get('perfis') or [session.get('perfil')]
    meu_setor = (session.get('setor') or '').strip()
    if meu_setor and 'admin' not in acessos and 'gerencia' in acessos:
        setores = [meu_setor]           # gerência de área vê só o seu setor
    else:
        setores = get_setores()         # admin/gerência geral veem todos
    return render_template('dashboard.html',
                           config_setores=setores,
                           turnos=get_turnos(),
                           meu_setor=meu_setor or get_setor_ativo())


@app.route('/admin/maquinas')
@perfil_obrigatorio('admin')
def admin_maquinas():
    db = Session()
    try:
        maquinas = db.query(Maquina).order_by(Maquina.setor, Maquina.ordem, Maquina.id).all()
        dados = [m.to_dict() for m in maquinas]
    finally:
        db.close()
    return render_template('admin_maquinas.html', maquinas=dados, setores=get_setores())


@app.route('/admin/usuarios')
@perfil_obrigatorio('admin')
def admin_usuarios():
    db = Session()
    try:
        usuarios = db.query(Usuario).order_by(Usuario.nome).all()
        dados = [u.to_dict() for u in usuarios]
    finally:
        db.close()
    return render_template('admin_usuarios.html', usuarios=dados, setores=get_setores())


@app.route('/admin/config')
@perfil_obrigatorio('admin')
def admin_config():
    return render_template('admin_config.html',
                           setores=get_setores(),
                           setor_ativo=get_setor_ativo(),
                           motivos=get_motivos_pausa(),
                           turnos=get_turnos(),
                           meta_pph_padrao=get_meta_pph_padrao(),
                           meta_oee=get_meta_oee())


@app.route('/admin/apontamentos')
@perfil_obrigatorio('admin')
def admin_apontamentos():
    return render_template('admin_apontamentos.html',
                           setores=get_setores(),
                           turnos=get_turnos(),
                           setor_ativo=get_setor_ativo())


# ─────────────────────────────────────────────────────────────────────────────
# API — utilidades
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/agora')
def api_agora():
    return jsonify({'iso': datetime.utcnow().isoformat() + 'Z'})


@app.route('/api/operadores')
@api_login_obrigatorio
def api_operadores():
    db = Session()
    try:
        us = db.query(Usuario).filter(Usuario.ativo.is_(True)).order_by(Usuario.nome).all()
        return jsonify([{'nome': u.nome, 'matricula': u.matricula or ''} for u in us])
    finally:
        db.close()


@app.route('/api/buscar_ordem/<path:ordem>')
@api_login_obrigatorio
def api_buscar_ordem(ordem):
    """Bipar/consultar a OP na lista mestra do SAP."""
    o = _norm_ordem(ordem)
    with _lista_lock:
        item = _lista_por_ordem.get(o)
    if item:
        return jsonify({'encontrado': True, **item})
    return jsonify({'encontrado': False, 'ordem': o})


@app.route('/api/lista_status')
@api_login_obrigatorio
def api_lista_status():
    with _lista_lock:
        st = dict(_lista_status)
    return jsonify(st)


@app.route('/api/admin/lista/recarregar', methods=['POST'])
@perfil_obrigatorio('admin')
def api_lista_recarregar():
    carregar_lista_mestre()
    with _lista_lock:
        return jsonify({'ok': True, **_lista_status})


# ─────────────────────────────────────────────────────────────────────────────
# API — grade de máquinas (estado atual do painel do operador)
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/maquinas')
@api_login_obrigatorio
def api_maquinas():
    setor = request.args.get('setor') or setor_do_usuario()
    agora = datetime.utcnow()
    db = Session()
    try:
        maquinas = (db.query(Maquina)
                    .filter(Maquina.setor == setor, Maquina.ativa.is_(True))
                    .order_by(Maquina.ordem, Maquina.id).all())
        ativos = (db.query(Apontamento)
                  .filter(Apontamento.setor == setor,
                          Apontamento.estado.in_(['produzindo', 'pausado']))
                  .all())
        por_maquina = {a.maquina_id: a for a in ativos}
        out = []
        for m in maquinas:
            d = m.to_dict()
            ap = por_maquina.get(m.id)
            d['apontamento'] = ap.to_dict(agora) if ap else None
            out.append(d)
        return jsonify({'setor': setor, 'maquinas': out,
                        'motivos': get_motivos_pausa(),
                        'agora': agora.isoformat() + 'Z'})
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# API — ciclo de vida do apontamento
# ─────────────────────────────────────────────────────────────────────────────
def _int(v, padrao=0):
    try:
        return int(float(str(v).replace(',', '.')))
    except (ValueError, TypeError):
        return padrao


def _float(v, padrao=0.0):
    try:
        return float(str(v).replace(',', '.'))
    except (ValueError, TypeError):
        return padrao


@app.route('/api/apontamento/iniciar', methods=['POST'])
@api_login_obrigatorio
def api_iniciar():
    d = request.get_json(force=True, silent=True) or {}
    maquina_id = _int(d.get('maquina_id'), 0)
    op = (d.get('op') or '').strip()          # opcional no início — a OP é bipada no final
    operador = (d.get('operador_nome') or session.get('nome') or '').strip()
    if not maquina_id:
        return jsonify({'ok': False, 'erro': 'Máquina inválida.'}), 400
    if not operador:
        return jsonify({'ok': False, 'erro': 'Informe o operador.'}), 400

    db = Session()
    try:
        m = db.query(Maquina).filter_by(id=maquina_id).first()
        if not m:
            return jsonify({'ok': False, 'erro': 'Máquina não encontrada.'}), 404
        # Já existe apontamento aberto nessa máquina?
        aberto = (db.query(Apontamento)
                  .filter(Apontamento.maquina_id == maquina_id,
                          Apontamento.estado.in_(['produzindo', 'pausado']))
                  .first())
        if aberto:
            return jsonify({'ok': False, 'erro': 'Já existe apontamento em andamento nesta máquina.'}), 409

        # Meta de peças/hora vigente: informada na abertura, senão a da máquina,
        # senão o padrão global. Congelada no apontamento para relatórios estáveis.
        meta_pph = _float(d.get('meta_pph'), 0) or float(m.meta_pph or 0) or get_meta_pph_padrao()
        ap = Apontamento(
            estado='produzindo', maquina_id=m.id, maquina_nome=m.nome, setor=m.setor,
            operador_nome=operador,
            operador_matricula=(d.get('operador_matricula') or session.get('matricula') or '').strip(),
            op=op, codigo=(d.get('codigo') or '').strip(),
            descricao=(d.get('descricao') or '').strip(),
            quantidade_prevista=_int(d.get('quantidade_prevista'), 0),
            meta_pph=meta_pph,
            inicio=datetime.utcnow(), pausas_json='[]', criado_em=datetime.utcnow())
        db.add(ap)
        db.commit()
        return jsonify({'ok': True, 'apontamento': ap.to_dict()})
    finally:
        db.close()


def _apontamento_aberto(db, maquina_id):
    return (db.query(Apontamento)
            .filter(Apontamento.maquina_id == maquina_id,
                    Apontamento.estado.in_(['produzindo', 'pausado']))
            .first())


@app.route('/api/apontamento/pausar', methods=['POST'])
@api_login_obrigatorio
def api_pausar():
    d = request.get_json(force=True, silent=True) or {}
    maquina_id = _int(d.get('maquina_id'), 0)
    motivo = (d.get('motivo') or '').strip() or 'Pausa'
    db = Session()
    try:
        ap = _apontamento_aberto(db, maquina_id)
        if not ap:
            return jsonify({'ok': False, 'erro': 'Nenhum apontamento aberto.'}), 404
        if ap.estado == 'pausado':
            return jsonify({'ok': False, 'erro': 'Já está pausado.'}), 409
        ap.estado = 'pausado'
        ap.pausado = 1
        ap.pausa_inicio = datetime.utcnow()
        ap.pausa_motivo = motivo
        db.commit()
        return jsonify({'ok': True, 'apontamento': ap.to_dict()})
    finally:
        db.close()


@app.route('/api/apontamento/retomar', methods=['POST'])
@api_login_obrigatorio
def api_retomar():
    d = request.get_json(force=True, silent=True) or {}
    maquina_id = _int(d.get('maquina_id'), 0)
    db = Session()
    try:
        ap = _apontamento_aberto(db, maquina_id)
        if not ap:
            return jsonify({'ok': False, 'erro': 'Nenhum apontamento aberto.'}), 404
        if ap.estado != 'pausado' or not ap.pausa_inicio:
            return jsonify({'ok': False, 'erro': 'Não está pausado.'}), 409
        agora = datetime.utcnow()
        dur = (agora - ap.pausa_inicio).total_seconds()
        ap.pausa_acumulada_seg = int((ap.pausa_acumulada_seg or 0) + dur)
        pausas = ap._pausas()
        pausas.append({
            'ini': (ap.pausa_inicio - timedelta(hours=FUSO_LOCAL_HORAS)).strftime('%d/%m %H:%M:%S'),
            'fim': (agora - timedelta(hours=FUSO_LOCAL_HORAS)).strftime('%d/%m %H:%M:%S'),
            'motivo': ap.pausa_motivo or 'Pausa',
            'seg': int(dur),
        })
        ap.pausas_json = json.dumps(pausas, ensure_ascii=False)
        ap.estado = 'produzindo'
        ap.pausado = 0
        ap.pausa_inicio = None
        ap.pausa_motivo = ''
        db.commit()
        return jsonify({'ok': True, 'apontamento': ap.to_dict()})
    finally:
        db.close()


@app.route('/api/apontamento/finalizar', methods=['POST'])
@api_login_obrigatorio
def api_finalizar():
    d = request.get_json(force=True, silent=True) or {}
    maquina_id = _int(d.get('maquina_id'), 0)
    db = Session()
    try:
        ap = _apontamento_aberto(db, maquina_id)
        if not ap:
            return jsonify({'ok': False, 'erro': 'Nenhum apontamento aberto.'}), 404
        agora = datetime.utcnow()
        # Se estava pausado, fecha a pausa em curso
        if ap.estado == 'pausado' and ap.pausa_inicio:
            dur = (agora - ap.pausa_inicio).total_seconds()
            ap.pausa_acumulada_seg = int((ap.pausa_acumulada_seg or 0) + dur)
            pausas = ap._pausas()
            pausas.append({
                'ini': (ap.pausa_inicio - timedelta(hours=FUSO_LOCAL_HORAS)).strftime('%d/%m %H:%M:%S'),
                'fim': (agora - timedelta(hours=FUSO_LOCAL_HORAS)).strftime('%d/%m %H:%M:%S'),
                'motivo': ap.pausa_motivo or 'Pausa', 'seg': int(dur),
            })
            ap.pausas_json = json.dumps(pausas, ensure_ascii=False)
            ap.pausa_inicio = None
            ap.pausa_motivo = ''
        ap.fim = agora
        ap.producao_seg = round(ap.produtivo_seg(agora), 1)
        # OP bipada no final: puxa OP e dados da lista mestra do SAP.
        op = (d.get('op') or '').strip()
        if not op:
            return jsonify({'ok': False, 'erro': 'Bipe ou informe a OP para finalizar.'}), 400
        ap.op = op
        if d.get('codigo') is not None:
            ap.codigo = (d.get('codigo') or '').strip()
        if d.get('descricao') is not None:
            ap.descricao = (d.get('descricao') or '').strip()
        if d.get('quantidade_prevista') is not None:
            ap.quantidade_prevista = _int(d.get('quantidade_prevista'), 0)
        ap.quantidade_produzida = _int(d.get('quantidade_produzida'), 0)
        ap.refugo = _int(d.get('refugo'), 0)
        if d.get('observacao') is not None:
            ap.observacao = (d.get('observacao') or '').strip()
        ap.estado = 'finalizado'
        ap.pausado = 0
        db.commit()
        return jsonify({'ok': True, 'apontamento': ap.to_dict()})
    finally:
        db.close()


@app.route('/api/apontamento/cancelar', methods=['POST'])
@api_login_obrigatorio
def api_cancelar():
    """Cancela (descarta) um apontamento aberto — usado em início por engano."""
    d = request.get_json(force=True, silent=True) or {}
    maquina_id = _int(d.get('maquina_id'), 0)
    db = Session()
    try:
        ap = _apontamento_aberto(db, maquina_id)
        if not ap:
            return jsonify({'ok': False, 'erro': 'Nenhum apontamento aberto.'}), 404
        db.delete(ap)
        db.commit()
        return jsonify({'ok': True})
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# API — Dashboard gerencial
# ─────────────────────────────────────────────────────────────────────────────
def _parse_data(s, padrao):
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return padrao


def _intervalo_utc(di, df):
    """Converte datas locais (00:00 de di até 23:59 de df) para limites UTC."""
    ini_local = datetime.combine(di, datetime.min.time())
    fim_local = datetime.combine(df, datetime.max.time())
    return (ini_local + timedelta(hours=FUSO_LOCAL_HORAS),
            fim_local + timedelta(hours=FUSO_LOCAL_HORAS))


@app.route('/api/dashboard/dados')
@api_login_obrigatorio
def api_dashboard():
    hoje = (datetime.utcnow() - timedelta(hours=FUSO_LOCAL_HORAS)).date()
    di = _parse_data(request.args.get('de'), hoje)
    df = _parse_data(request.args.get('ate'), hoje)
    setor, setor_rotulo = resolver_setor(request.args.get('setor'))
    turno = (request.args.get('turno') or '').strip()
    filtro_maq = (request.args.get('maquina') or '').strip()
    ini_utc, fim_utc = _intervalo_utc(di, df)

    db = Session()
    try:
        q = (db.query(Apontamento)
             .filter(Apontamento.estado == 'finalizado',
                     Apontamento.inicio >= ini_utc,
                     Apontamento.inicio <= fim_utc)
             .order_by(Apontamento.fim.desc()))
        if setor:
            q = q.filter(Apontamento.setor == setor)
        finalizados = q.all()
        if turno:
            finalizados = [a for a in finalizados
                           if turno_de(a.inicio - timedelta(hours=FUSO_LOCAL_HORAS)) == turno]
        if filtro_maq:
            finalizados = [a for a in finalizados if a.maquina_nome == filtro_maq]

        # Em andamento (agora)
        qa = (db.query(Apontamento)
              .filter(Apontamento.estado.in_(['produzindo', 'pausado'])))
        if setor:
            qa = qa.filter(Apontamento.setor == setor)
        em_andamento = qa.all()
        if filtro_maq:
            em_andamento = [a for a in em_andamento if a.maquina_nome == filtro_maq]

        # Máquinas (para o filtro de máquina no painel)
        qm = db.query(Maquina).filter(Maquina.ativa.is_(True))
        if setor:
            qm = qm.filter(Maquina.setor == setor)
        maquinas_setor = [m.nome for m in qm.order_by(Maquina.setor, Maquina.ordem, Maquina.id).all()]

        meta_padrao = get_meta_pph_padrao()
        meta_oee = get_meta_oee()

        total_prod_seg = sum(a.producao_seg or 0 for a in finalizados)
        total_pausa_seg = sum(a.pausa_total_seg() for a in finalizados)
        total_pecas = sum(a.quantidade_produzida or 0 for a in finalizados)
        total_refugo = sum(a.refugo or 0 for a in finalizados)
        # produção teórica (p/ desempenho): Σ meta_efetiva * horas produtivas
        total_teorico = sum(a.meta_efetiva(meta_padrao) * ((a.producao_seg or 0) / 3600)
                            for a in finalizados)
        n = len(finalizados)

        eficiencia = 0.0
        base = total_prod_seg + total_pausa_seg
        if base > 0:
            eficiencia = round(total_prod_seg / base * 100, 1)

        # OEE agregado do período (disponibilidade x desempenho x qualidade)
        oee_geral = calc_oee(total_prod_seg, total_pausa_seg, total_pecas,
                             total_refugo, 0)
        # desempenho agregado usa a produção teórica somada
        desemp_geral = min(total_pecas / total_teorico, 1.0) if total_teorico > 0 else None
        disp_geral = oee_geral['disponibilidade']
        qual_geral = oee_geral['qualidade']
        if None in (disp_geral, desemp_geral, qual_geral):
            oee_total = None
        else:
            oee_total = disp_geral * desemp_geral * qual_geral

        def _novo_grupo(k):
            return {'nome': k, 'pecas': 0, 'prod_seg': 0.0, 'pausa_seg': 0.0,
                    'ops': 0, 'refugo': 0, 'teorico': 0.0}

        def _acum(r, a):
            r['pecas'] += a.quantidade_produzida or 0
            r['prod_seg'] += a.producao_seg or 0
            r['pausa_seg'] += a.pausa_total_seg()
            r['refugo'] += a.refugo or 0
            r['teorico'] += a.meta_efetiva(meta_padrao) * ((a.producao_seg or 0) / 3600)
            r['ops'] += 1

        # Por operador
        por_op = {}
        for a in finalizados:
            _acum(por_op.setdefault(a.operador_nome or '—',
                                    _novo_grupo(a.operador_nome or '—')), a)

        # Por máquina
        por_maq = {}
        for a in finalizados:
            _acum(por_maq.setdefault(a.maquina_nome or '—',
                                     _novo_grupo(a.maquina_nome or '—')), a)

        # Motivos de pausa (soma dos segundos)
        motivos = {}
        for a in finalizados:
            for p in a._pausas():
                mv = p.get('motivo') or 'Pausa'
                motivos[mv] = motivos.get(mv, 0) + int(p.get('seg') or 0)

        # Produção por dia (peças + horas produtivas)
        por_dia = {}
        for a in finalizados:
            dloc = (a.inicio - timedelta(hours=FUSO_LOCAL_HORAS)).date().isoformat() if a.inicio else '—'
            r = por_dia.setdefault(dloc, {'dia': dloc, 'pecas': 0, 'prod_h': 0.0})
            r['pecas'] += a.quantidade_produzida or 0
            r['prod_h'] += (a.producao_seg or 0) / 3600

        def arr(d, chave_ord):
            lst = list(d.values())
            for r in lst:
                if 'prod_seg' in r:
                    r['prod_min'] = round(r['prod_seg'] / 60, 1)
                    r['prod_h'] = round(r['prod_seg'] / 3600, 2)
                    r['pausa_min'] = round(r['pausa_seg'] / 60, 1)
                    r['pph'] = round(r['pecas'] / (r['prod_seg'] / 3600), 1) if r['prod_seg'] > 0 and r['pecas'] else 0.0
                    disp = (r['prod_seg'] / (r['prod_seg'] + r['pausa_seg'])) if (r['prod_seg'] + r['pausa_seg']) > 0 else None
                    desemp = min(r['pecas'] / r['teorico'], 1.0) if r['teorico'] > 0 else None
                    qual = ((r['pecas'] - r['refugo']) / r['pecas']) if r['pecas'] > 0 else None
                    oee = disp * desemp * qual if None not in (disp, desemp, qual) else None
                    r['disponibilidade'] = _pct(disp)
                    r['desempenho'] = _pct(desemp)
                    r['qualidade'] = _pct(qual)
                    r['oee'] = _pct(oee)
            lst.sort(key=lambda x: x.get(chave_ord, 0) or 0, reverse=True)
            return lst

        por_dia_lst = list(por_dia.values())
        for r in por_dia_lst:
            r['prod_h'] = round(r['prod_h'], 2)
        por_dia_lst.sort(key=lambda x: x['dia'])

        motivos_lst = [{'motivo': k, 'seg': v, 'min': round(v / 60, 1)}
                       for k, v in motivos.items()]
        motivos_lst.sort(key=lambda x: x['seg'], reverse=True)

        agora = datetime.utcnow()
        return jsonify({
            'setor': setor_rotulo, 'de': di.isoformat(), 'ate': df.isoformat(),
            'turno': turno, 'turnos': [t['nome'] for t in get_turnos()],
            'maquina': filtro_maq, 'maquinas': maquinas_setor,
            'kpis': {
                'ops': n,
                'pecas': total_pecas,
                'refugo': total_refugo,
                'prod_h': round(total_prod_seg / 3600, 2),
                'pausa_h': round(total_pausa_seg / 3600, 2),
                'eficiencia': eficiencia,
                'pph': round(total_pecas / (total_prod_seg / 3600), 1) if total_prod_seg > 0 and total_pecas else 0.0,
                'em_andamento': len(em_andamento),
                # OEE e pilares (percentuais; null quando não há meta cadastrada)
                'disponibilidade': _pct(disp_geral),
                'desempenho': _pct(desemp_geral),
                'qualidade': _pct(qual_geral),
                'oee': _pct(oee_total),
                'meta_oee': meta_oee,
            },
            'meta_oee': meta_oee,
            'por_operador': arr(por_op, 'pecas'),
            'por_maquina': arr(por_maq, 'pecas'),
            'motivos_pausa': motivos_lst,
            'por_dia': por_dia_lst,
            'em_andamento': [a.to_dict(agora, meta_padrao) for a in em_andamento],
            'ultimos': [a.to_dict(agora, meta_padrao) for a in finalizados[:40]],
        })
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# API — Admin: máquinas
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/admin/maquina/salvar', methods=['POST'])
@perfil_obrigatorio('admin')
def api_maquina_salvar():
    d = request.get_json(force=True, silent=True) or {}
    nome = (d.get('nome') or '').strip()
    if not nome:
        return jsonify({'ok': False, 'erro': 'Informe o nome da máquina.'}), 400
    db = Session()
    try:
        mid = _int(d.get('id'), 0)
        if mid:
            m = db.query(Maquina).filter_by(id=mid).first()
            if not m:
                return jsonify({'ok': False, 'erro': 'Máquina não encontrada.'}), 404
        else:
            m = Maquina()
            db.add(m)
        m.nome = nome
        m.codigo = (d.get('codigo') or '').strip()
        m.setor = (d.get('setor') or 'Dobra').strip()
        m.ativa = bool(d.get('ativa', True))
        m.ordem = _int(d.get('ordem'), 0)
        m.meta_pph = max(0.0, _float(d.get('meta_pph'), 0))
        db.commit()
        return jsonify({'ok': True, 'maquina': m.to_dict()})
    finally:
        db.close()


@app.route('/api/admin/maquina/excluir', methods=['POST'])
@perfil_obrigatorio('admin')
def api_maquina_excluir():
    d = request.get_json(force=True, silent=True) or {}
    mid = _int(d.get('id'), 0)
    db = Session()
    try:
        m = db.query(Maquina).filter_by(id=mid).first()
        if not m:
            return jsonify({'ok': False, 'erro': 'Máquina não encontrada.'}), 404
        aberto = _apontamento_aberto(db, mid)
        if aberto:
            return jsonify({'ok': False, 'erro': 'Há apontamento em andamento nesta máquina.'}), 409
        db.delete(m)
        db.commit()
        return jsonify({'ok': True})
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# API — Admin: usuários
# ─────────────────────────────────────────────────────────────────────────────
PERFIS_VALIDOS = ['operador', 'gerencia', 'admin']


@app.route('/api/admin/usuario/salvar', methods=['POST'])
@perfil_obrigatorio('admin')
def api_usuario_salvar():
    d = request.get_json(force=True, silent=True) or {}
    login_ = (d.get('login') or '').strip().lower()
    nome = (d.get('nome') or '').strip()
    perfis = [p for p in (d.get('perfis') or []) if p in PERFIS_VALIDOS]
    if not login_ or not nome:
        return jsonify({'ok': False, 'erro': 'Login e nome são obrigatórios.'}), 400
    if not perfis:
        return jsonify({'ok': False, 'erro': 'Selecione ao menos um acesso.'}), 400
    db = Session()
    try:
        uid = _int(d.get('id'), 0)
        if uid:
            u = db.query(Usuario).filter_by(id=uid).first()
            if not u:
                return jsonify({'ok': False, 'erro': 'Usuário não encontrado.'}), 404
        else:
            if db.query(Usuario).filter_by(login=login_).first():
                return jsonify({'ok': False, 'erro': 'Já existe um usuário com esse login.'}), 409
            u = Usuario(login=login_)
            db.add(u)
        # Se mudou o login num usuário existente, checa duplicidade
        if uid and login_ != u.login and db.query(Usuario).filter_by(login=login_).first():
            return jsonify({'ok': False, 'erro': 'Já existe um usuário com esse login.'}), 409
        u.login = login_
        u.nome = nome
        u.matricula = (d.get('matricula') or '').strip()
        u.setor = (d.get('setor') or '').strip()
        u.perfis = ','.join(perfis)
        u.perfil = 'admin' if 'admin' in perfis else ('gerencia' if 'gerencia' in perfis else 'operador')
        u.ativo = bool(d.get('ativo', True))
        senha = (d.get('senha') or '').strip()
        if senha:
            u.senha_hash = generate_password_hash(senha)
        elif not uid:
            return jsonify({'ok': False, 'erro': 'Defina uma senha para o novo usuário.'}), 400
        db.commit()
        return jsonify({'ok': True, 'usuario': u.to_dict()})
    finally:
        db.close()


@app.route('/api/admin/usuario/excluir', methods=['POST'])
@perfil_obrigatorio('admin')
def api_usuario_excluir():
    d = request.get_json(force=True, silent=True) or {}
    uid = _int(d.get('id'), 0)
    db = Session()
    try:
        u = db.query(Usuario).filter_by(id=uid).first()
        if not u:
            return jsonify({'ok': False, 'erro': 'Usuário não encontrado.'}), 404
        if u.login == 'admin':
            return jsonify({'ok': False, 'erro': 'O usuário admin não pode ser excluído.'}), 400
        if u.login == session.get('usuario'):
            return jsonify({'ok': False, 'erro': 'Você não pode excluir a si mesmo.'}), 400
        db.delete(u)
        db.commit()
        return jsonify({'ok': True})
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# API — Admin: configurações (setores, setor ativo, motivos de pausa)
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/admin/config/salvar', methods=['POST'])
@perfil_obrigatorio('admin')
def api_config_salvar():
    d = request.get_json(force=True, silent=True) or {}
    setores = [s.strip() for s in (d.get('setores') or []) if s and s.strip()]
    motivos = [m.strip() for m in (d.get('motivos_pausa') or []) if m and m.strip()]
    setor_ativo = (d.get('setor_ativo') or '').strip()
    if setores:
        cfg_set('setores', setores)
    if motivos:
        cfg_set('motivos_pausa', motivos)
    if setor_ativo:
        cfg_set('setor_ativo', setor_ativo)
    if 'meta_pph_padrao' in d:
        cfg_set('meta_pph_padrao', max(0.0, _float(d.get('meta_pph_padrao'), 0)))
    if 'meta_oee' in d:
        cfg_set('meta_oee', min(100.0, max(0.0, _float(d.get('meta_oee'), META_OEE_PADRAO))))
    if isinstance(d.get('turnos'), list):
        turnos = []
        for t in d['turnos']:
            if not isinstance(t, dict):
                continue
            nome = (t.get('nome') or '').strip()
            if not nome:
                continue
            turnos.append({'nome': nome,
                           'inicio': (t.get('inicio') or '00:00').strip(),
                           'fim': (t.get('fim') or '00:00').strip()})
        if turnos:
            cfg_set('turnos', turnos)
    return jsonify({'ok': True, 'setores': get_setores(),
                    'setor_ativo': get_setor_ativo(),
                    'motivos_pausa': get_motivos_pausa(),
                    'meta_pph_padrao': get_meta_pph_padrao(),
                    'meta_oee': get_meta_oee(),
                    'turnos': get_turnos()})


# ─────────────────────────────────────────────────────────────────────────────
# API — Exportação Excel
# ─────────────────────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill('solid', fgColor='1668C0')
_HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
_BORDER = Border(*[Side(style='thin', color='D0D7E2')] * 4)
_CENTER = Alignment(horizontal='center', vertical='center')


def _seg_hms(seg):
    seg = int(seg or 0)
    h, r = divmod(seg, 3600)
    m, s = divmod(r, 60)
    return f'{h:02d}:{m:02d}:{s:02d}'


@app.route('/api/download/apontamentos')
@perfil_obrigatorio('gerencia', 'admin')
def download_apontamentos():
    hoje = (datetime.utcnow() - timedelta(hours=FUSO_LOCAL_HORAS)).date()
    di = _parse_data(request.args.get('de'), hoje)
    df = _parse_data(request.args.get('ate'), hoje)
    setor, setor_rotulo = resolver_setor(request.args.get('setor'))
    turno = (request.args.get('turno') or '').strip()
    filtro_maq = (request.args.get('maquina') or '').strip()
    ini_utc, fim_utc = _intervalo_utc(di, df)

    db = Session()
    try:
        q = (db.query(Apontamento)
             .filter(Apontamento.estado == 'finalizado',
                     Apontamento.inicio >= ini_utc,
                     Apontamento.inicio <= fim_utc)
             .order_by(Apontamento.inicio))
        if setor:
            q = q.filter(Apontamento.setor == setor)
        aps = q.all()
        if turno:
            aps = [a for a in aps
                   if turno_de(a.inicio - timedelta(hours=FUSO_LOCAL_HORAS)) == turno]
        if filtro_maq:
            aps = [a for a in aps if a.maquina_nome == filtro_maq]
        wb = Workbook()
        ws = wb.active
        ws.title = 'Apontamentos'
        meta_padrao = get_meta_pph_padrao()
        cols = ['Data', 'Turno', 'Setor', 'Máquina', 'Operador', 'Matrícula', 'OP', 'Código',
                'Descrição', 'Qtd prevista', 'Qtd produzida', 'Refugo',
                'Início', 'Fim', 'Tempo produtivo', 'Pausa total',
                'Nº pausas', 'Meta pç/h', 'Peças/hora',
                'Disponibilidade %', 'Desempenho %', 'Qualidade %', 'OEE %',
                'Observação']
        ws.append(cols)
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
            cell.border = _BORDER

        def _pex(x):
            return '' if x is None else x

        for a in aps:
            dloc = (a.inicio - timedelta(hours=FUSO_LOCAL_HORAS)) if a.inicio else None
            fimloc = (a.fim - timedelta(hours=FUSO_LOCAL_HORAS)) if a.fim else None
            prod = a.producao_seg or 0
            pph = round((a.quantidade_produzida or 0) / (prod / 3600), 1) if prod > 0 and a.quantidade_produzida else 0
            o = a.oee(meta_padrao=meta_padrao)
            ws.append([
                dloc.strftime('%d/%m/%Y') if dloc else '',
                turno_de(dloc) if dloc else '',
                a.setor or '',
                a.maquina_nome, a.operador_nome, a.operador_matricula or '',
                a.op or '', a.codigo or '', a.descricao or '',
                a.quantidade_prevista or 0, a.quantidade_produzida or 0, a.refugo or 0,
                dloc.strftime('%H:%M:%S') if dloc else '',
                fimloc.strftime('%H:%M:%S') if fimloc else '',
                _seg_hms(prod), _seg_hms(a.pausa_total_seg()),
                len(a._pausas()), round(a.meta_efetiva(meta_padrao), 2), pph,
                _pex(_pct(o['disponibilidade'])), _pex(_pct(o['desempenho'])),
                _pex(_pct(o['qualidade'])), _pex(_pct(o['oee'])),
                a.observacao or '',
            ])
        larguras = [12, 10, 14, 16, 20, 12, 12, 14, 30, 12, 13, 10, 10, 10, 15, 13,
                    10, 10, 11, 16, 15, 14, 10, 30]
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        ws.freeze_panes = 'A2'
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        _sl = (setor or 'todos_setores').replace('/', '-').replace(' ', '_')
        nome = f'apontamentos_{_sl}_{di.isoformat()}_a_{df.isoformat()}.xlsx'
        return send_file(bio, as_attachment=True, download_name=nome,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# API — Admin: editar / excluir apontamentos (com auditoria)
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/admin/apontamentos')
@perfil_obrigatorio('admin')
def api_admin_apontamentos():
    hoje = (datetime.utcnow() - timedelta(hours=FUSO_LOCAL_HORAS)).date()
    di = _parse_data(request.args.get('de'), hoje)
    df = _parse_data(request.args.get('ate'), hoje)
    setor = request.args.get('setor') or get_setor_ativo()
    turno = (request.args.get('turno') or '').strip()
    ini_utc, fim_utc = _intervalo_utc(di, df)
    agora = datetime.utcnow()
    meta_padrao = get_meta_pph_padrao()
    db = Session()
    try:
        aps = (db.query(Apontamento)
               .filter(Apontamento.estado == 'finalizado',
                       Apontamento.setor == setor,
                       Apontamento.inicio >= ini_utc,
                       Apontamento.inicio <= fim_utc)
               .order_by(Apontamento.inicio.desc()).all())
        if turno:
            aps = [a for a in aps
                   if turno_de(a.inicio - timedelta(hours=FUSO_LOCAL_HORAS)) == turno]
        return jsonify({'setor': setor, 'turno': turno,
                        'de': di.isoformat(), 'ate': df.isoformat(),
                        'apontamentos': [a.to_dict(agora, meta_padrao) for a in aps]})
    finally:
        db.close()


_CAMPOS_EDITAVEIS = ['op', 'codigo', 'descricao', 'operador_nome', 'operador_matricula',
                     'quantidade_prevista', 'quantidade_produzida', 'refugo',
                     'meta_pph', 'observacao']


def _snapshot(ap):
    return {c: getattr(ap, c) for c in _CAMPOS_EDITAVEIS}


@app.route('/api/admin/apontamento/salvar', methods=['POST'])
@perfil_obrigatorio('admin')
def api_apontamento_salvar():
    d = request.get_json(force=True, silent=True) or {}
    aid = _int(d.get('id'), 0)
    if not aid:
        return jsonify({'ok': False, 'erro': 'Apontamento inválido.'}), 400
    db = Session()
    try:
        ap = db.query(Apontamento).filter_by(id=aid).first()
        if not ap:
            return jsonify({'ok': False, 'erro': 'Apontamento não encontrado.'}), 404
        antes = _snapshot(ap)
        # Texto
        for campo in ['op', 'codigo', 'descricao', 'operador_nome', 'operador_matricula', 'observacao']:
            if campo in d:
                setattr(ap, campo, (d.get(campo) or '').strip())
        # Números
        if 'quantidade_prevista' in d:
            ap.quantidade_prevista = max(0, _int(d.get('quantidade_prevista'), 0))
        if 'quantidade_produzida' in d:
            ap.quantidade_produzida = max(0, _int(d.get('quantidade_produzida'), 0))
        if 'refugo' in d:
            ap.refugo = max(0, _int(d.get('refugo'), 0))
        if 'meta_pph' in d:
            ap.meta_pph = max(0.0, _float(d.get('meta_pph'), 0))
        db.add(ApontamentoLog(
            apontamento_id=ap.id, quando=datetime.utcnow(),
            usuario=session.get('nome') or session.get('usuario') or '',
            acao='editar',
            antes_json=json.dumps(antes, ensure_ascii=False, default=str),
            depois_json=json.dumps(_snapshot(ap), ensure_ascii=False, default=str)))
        db.commit()
        return jsonify({'ok': True, 'apontamento': ap.to_dict(meta_padrao=get_meta_pph_padrao())})
    finally:
        db.close()


@app.route('/api/admin/apontamento/excluir', methods=['POST'])
@perfil_obrigatorio('admin')
def api_apontamento_excluir():
    d = request.get_json(force=True, silent=True) or {}
    aid = _int(d.get('id'), 0)
    if not aid:
        return jsonify({'ok': False, 'erro': 'Apontamento inválido.'}), 400
    db = Session()
    try:
        ap = db.query(Apontamento).filter_by(id=aid).first()
        if not ap:
            return jsonify({'ok': False, 'erro': 'Apontamento não encontrado.'}), 404
        db.add(ApontamentoLog(
            apontamento_id=ap.id, quando=datetime.utcnow(),
            usuario=session.get('nome') or session.get('usuario') or '',
            acao='excluir',
            antes_json=json.dumps(_snapshot(ap), ensure_ascii=False, default=str),
            depois_json=''))
        db.delete(ap)
        db.commit()
        return jsonify({'ok': True})
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# API — Admin: backup e restauração (base completa do site)
# ─────────────────────────────────────────────────────────────────────────────
_BACKUP_MODELOS = [
    ('config', Config),
    ('usuarios', Usuario),
    ('maquinas', Maquina),
    ('apontamentos', Apontamento),
    ('logs', ApontamentoLog),
]


def _dump_row(obj):
    d = {}
    for col in obj.__table__.columns:
        v = getattr(obj, col.name)
        if isinstance(v, datetime):
            v = v.isoformat()
        d[col.name] = v
    return d


def _load_row(Model, d):
    cols = {c.name: c for c in Model.__table__.columns}
    kwargs = {}
    for k, v in (d or {}).items():
        col = cols.get(k)
        if col is None:
            continue
        if v is not None and isinstance(col.type, DateTime):
            try:
                v = datetime.fromisoformat(str(v).replace('Z', ''))
            except (ValueError, TypeError):
                v = None
        kwargs[k] = v
    return Model(**kwargs)


def _reset_sequences(db):
    """No PostgreSQL, ressincroniza as sequences de id após restaurar ids explícitos."""
    if engine.dialect.name != 'postgresql':
        return
    for _, Model in _BACKUP_MODELOS:
        tabela = Model.__tablename__
        db.execute(text(
            "SELECT setval(pg_get_serial_sequence(:t, 'id'), "
            "(SELECT COALESCE(MAX(id), 1) FROM " + tabela + "))"
        ), {'t': tabela})


@app.route('/api/admin/backup')
@perfil_obrigatorio('admin')
def api_backup():
    db = Session()
    try:
        data = {'version': 1, 'gerado_em': datetime.utcnow().isoformat() + 'Z'}
        for chave, Model in _BACKUP_MODELOS:
            data[chave] = [_dump_row(x) for x in db.query(Model).all()]
        bio = io.BytesIO(json.dumps(data, ensure_ascii=False, default=str).encode('utf-8'))
        bio.seek(0)
        nome = f"backup_producao_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.json"
        return send_file(bio, as_attachment=True, download_name=nome,
                         mimetype='application/json')
    finally:
        db.close()


@app.route('/api/admin/restore', methods=['POST'])
@perfil_obrigatorio('admin')
def api_restore():
    # Aceita arquivo (multipart 'arquivo') ou JSON no corpo.
    payload = None
    arq = request.files.get('arquivo')
    if arq is not None:
        try:
            payload = json.loads(arq.read().decode('utf-8'))
        except (ValueError, UnicodeDecodeError):
            return jsonify({'ok': False, 'erro': 'Arquivo de backup inválido.'}), 400
    else:
        payload = request.get_json(force=True, silent=True)
    if not isinstance(payload, dict) or 'usuarios' not in payload or 'maquinas' not in payload:
        return jsonify({'ok': False, 'erro': 'Backup inválido ou incompleto.'}), 400

    db = Session()
    try:
        # Apaga tudo (filhos primeiro por causa das referências).
        for chave, Model in reversed(_BACKUP_MODELOS):
            db.query(Model).delete()
        db.commit()
        # Reinsere preservando os ids.
        contagem = {}
        for chave, Model in _BACKUP_MODELOS:
            linhas = payload.get(chave, []) or []
            for d in linhas:
                db.add(_load_row(Model, d))
            contagem[chave] = len(linhas)
        db.commit()
        _reset_sequences(db)
        db.commit()
        return jsonify({'ok': True, 'contagem': contagem})
    except Exception as e:  # noqa: BLE001 — devolve o motivo ao admin
        db.rollback()
        return jsonify({'ok': False, 'erro': 'Falha ao restaurar: ' + str(e)}), 500
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# API — Admin: reset (zera apontamentos) — proteção por confirmação
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/admin/reset', methods=['POST'])
@perfil_obrigatorio('admin')
def api_reset():
    d = request.get_json(force=True, silent=True) or {}
    if (d.get('confirmar') or '').strip().upper() != 'APAGAR':
        return jsonify({'ok': False, 'erro': 'Digite APAGAR para confirmar.'}), 400
    db = Session()
    try:
        n = db.query(Apontamento).delete()
        db.query(ApontamentoLog).delete()
        db.commit()
        return jsonify({'ok': True, 'apagados': n})
    finally:
        db.close()


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
